#!/usr/bin/env python3
"""
生成示例学员 Excel 文件用于测试
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_sample_excel():
    """创建示例 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "学员数据"

    # 设置表头
    headers = ["姓名", "性别", "年龄", "学员类型", "优先级"]
    ws.append(headers)

    # 设置表头格式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 示例数据
    sample_data = [
        ["法悟", "M", 45, "monk", 1],
        ["李梅", "F", 38, "old_student", 2],
        ["张三", "M", 25, "new_student", 3],
        ["王五", "M", 42, "old_student", 2],
        ["陈女", "F", 30, "new_student", 3],
        ["林道", "M", 55, "monk", 1],
        ["刘英", "F", 28, "new_student", 3],
        ["孙慧", "F", 48, "old_student", 2],
        ["赵刚", "M", 35, "old_student", 2],
        ["周翔", "M", 26, "new_student", 3],
    ]

    # 添加数据
    for row_data in sample_data:
        ws.append(row_data)

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10

    # 设置数据行对齐
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 保存文件
    output_file = "sample_students.xlsx"
    wb.save(output_file)
    print(f"✓ 示例文件已生成: {output_file}")
    print(f"  文件位置: {os.path.abspath(output_file)}")
    print("\n说明：")
    print("- 第1列：姓名（必填）")
    print("- 第2列：性别（必填，M 男 / F 女）")
    print("- 第3列：年龄（必填，正整数）")
    print("- 第4列：学员类型（必填）")
    print("    • monk 表示法师")
    print("    • old_student 表示旧生")
    print("    • new_student 表示新生")
    print("- 第5列：优先级（可选，数字）")

if __name__ == "__main__":
    try:
        create_sample_excel()
    except ImportError:
        print("错误：需要安装 openpyxl")
        print("请运行: pip install openpyxl")
